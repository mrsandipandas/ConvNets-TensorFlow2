import tensorflow as tf
import argparse
import utils
import numpy as np
import os
from openpyxl import Workbook
from openpyxl import load_workbook

# import wandb

def inferenceCheck(excel_path, model_name, dataset):
    checked = False
    wb = load_workbook(excel_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            if cell.value is not None:
                if (model_name == cell.value):
                    print(model_name, 'model has been inferenced already for', dataset)
                    checked = True
                    break
    return checked

os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'

parser = argparse.ArgumentParser()
parser.add_argument('--nets', type=str, required=True)
parser.add_argument('--batch_size', type=int, default=32)
parser.add_argument('--lr', type=float, default=0.01)
parser.add_argument('--epochs', type=int, default=50)
parser.add_argument('--ops', type=str, default='train')
parser.add_argument('--saved_model', type=str, default='./saved_model')
parser.add_argument('--dataset', type=str, default='cifar10') # cifar10 | cifar100
args = parser.parse_args()

print(args)

saved_model_path = '{}/{}/{}.h5'
path = saved_model_path.format(args.saved_model, args.dataset, args.nets)
# wandb.init(project="conv-nets", name=args.nets.lower())

dataset_type = args.dataset
if dataset_type == 'cifar10':
    dataset = tf.keras.datasets.cifar10
    num_classes = 10
elif dataset_type == 'cifar100':
    dataset = tf.keras.datasets.cifar100
    num_classes = 100
else:
    raise ValueError('Please use cifar10 or cifar100 dataset')

(x_train, y_train), (x_test, y_test) = dataset.load_data()
x_train, x_test = x_train / 255.0, x_test / 255.0

print('Unique training labels : ',np.unique(y_train))

if args.ops == 'train':
    # tf.debugging.set_log_device_placement(True)
    # These 2 lines below use all GPUs present in the host
    strategy = tf.distribute.MirroredStrategy()
    with strategy.scope():
        model = utils.choose_nets(args.nets, num_classes)
        
        model.compile(optimizer='adam',
                    loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
                    metrics=['accuracy'])

        model.fit(x_train, y_train,
                validation_split=0.1,
                epochs=args.epochs,
                batch_size=args.batch_size)

        _, baseline_model_accuracy = model.evaluate(x_test, y_test, verbose=0)
        print('Baseline test accuracy:', baseline_model_accuracy)
        model.save_weights(path)

if args.ops == 'test':
    saved_table_path = '{}/{}/{}.xlsx'
    table_path = saved_table_path.format(args.saved_model, args.dataset, args.dataset)

    checked = inferenceCheck(table_path, args.nets, args.dataset)

    if not checked:
        print("Trained model is present at: ", path)
        model = utils.choose_nets(args.nets, num_classes)

        model.compile(optimizer='adam',
                    loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
                    metrics=['accuracy'])    
        
        # This line is needed to initialize before loading the weights
        model.train_on_batch(x_train[:1], y_train[:1])
        model.load_weights(path)
        print("Loaded model details:", args.nets, "for", args.dataset, "dataset")
        print(model.summary())
        
        # Trained model
        train_loss, train_acc = model.evaluate(x_train, y_train, verbose=2)
        print("-" * 50)
        print('Restored model, training dataset accuracy: {:5.2f}%'.format(100*train_acc))
        
        print("-" * 50)

        # Test model
        test_loss, test_acc = model.evaluate(x_test, y_test, verbose=2)
        print("-" * 50)
        print('Restored model, testing dataset accuracy: {:5.2f}%'.format(100*test_acc))

        # Put the results in worksheet
        print('The path to the excel comparision file is:', table_path)
        wb = load_workbook(table_path)
        ws = wb.active
        ws.append([args.nets, train_acc, test_acc])
        wb.save(table_path)


        #for layer in model.layers: 
            #print(layer.get_weights())

        last_layer_weights = model.layers[-1].get_weights()[0]
        last_layer_biases  = model.layers[-1].get_weights()[1]
        print ('The last layer weight shape is: ', np.shape(last_layer_weights))
        print ('The last layer bias shape is: ', np.shape(last_layer_biases))

        # wandb.log({
        #     "TrainLoss": train_loss.result(),
        #     "TestLoss": test_loss.result(),
        #     "TrainAcc": train_accuracy.result()*100,
        #     "TestAcc": test_accuracy.result()*100
        # })
